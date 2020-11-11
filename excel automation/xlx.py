import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def xlautomate():
    
    filename = input("enter your filename :> ")
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    for row in range(2, sheet.max_row+1):
        cell = sheet.cell(row, 3)
        cell_corrected_value = cell.value * 0.9
        cell_corrected_position = sheet.cell(row,4)
        cell_corrected_position.value = cell_corrected_value
        
    ref_value = Reference(sheet, min_row=2, max_row = sheet.max_row, min_col=4, max_col=4)
    b = BarChart()
    b.add_data(ref_value)
    sheet.add_chart(b, "e2")

    wb.save("transaction2.xlsx")


xlautomate() 

